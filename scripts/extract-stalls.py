"""
Generate the canonical CTH stall inventory + geometry from the official Excel
site plan (SDP2026). Output: public/stalls.json — the source of truth for the
interactive 2D map, slot counters, and allocation.

Numbering is preserved EXACTLY as drawn (BS13-36, FT01-47, etc.). We only strip
the wrap-whitespace inside a label ("TS 6" -> "TS6"); we never renumber.

Each stall's rectangle = the merged-cell block that holds its label, in grid
units (col,row,w,h). The SVG map scales grid units to pixels 1:1 with the plan.
"""
import json
import os
import re

import openpyxl

SRC = "/Users/milaaj/Downloads/CTH Template layout 2026 (1).xlsx"
OUT = os.path.join(os.path.dirname(__file__), "..", "public", "stalls.json")
SHEET = "SDP2026 for website V2"

STALL_RE = re.compile(r"^(FT|FS|TS|BS)\s*0*(\d+)$", re.I)
ZONE_WORDS = ("ENTRANCE", "EXIT", "CARNIVAL", "JUMP", "SALAAH", "GENERATOR",
              "TICKETS", "PETTING")


def main():
    wb = openpyxl.load_workbook(SRC)
    ws = wb[SHEET]
    merges = list(ws.merged_cells.ranges)

    def merge_at(r, c):
        for m in merges:
            if m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col:
                return (m.min_col, m.min_row, m.max_col - m.min_col + 1,
                        m.max_row - m.min_row + 1)
        return (c, r, 1, 1)

    stalls, zones = [], []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            raw = v.strip()
            flat = re.sub(r"\s+", "", raw)
            m = STALL_RE.match(flat)
            if m:
                col, rw, w, h = merge_at(cell.row, cell.column)
                stalls.append({
                    "code": flat.upper(),                 # exact, e.g. FT01, FS1, TS6, BS13
                    "type": m.group(1).upper(),
                    "num": int(m.group(2)),
                    "col": col, "row": rw, "w": w, "h": h,
                })
            elif any(word in raw.upper() for word in ZONE_WORDS) and len(raw) < 40:
                col, rw, w, h = merge_at(cell.row, cell.column)
                zones.append({"label": " ".join(raw.split()),
                              "col": col, "row": rw, "w": w, "h": h})

    # de-dupe stalls by code (keep first), guard against accidental doubles
    seen, uniq = set(), []
    for s in sorted(stalls, key=lambda x: (x["type"], x["num"])):
        if s["code"] in seen:
            continue
        seen.add(s["code"])
        uniq.append(s)

    caps = {}
    for s in uniq:
        caps[s["type"]] = caps.get(s["type"], 0) + 1

    out = {
        "grid": {"cols": ws.max_column, "rows": ws.max_row},
        "capacity": caps,
        "types": {
            "FT": {"label": "Food Truck", "color": "#f97316"},
            "FS": {"label": "Full Space", "color": "#a855f7"},
            "TS": {"label": "Table Space", "color": "#eab308"},
            "BS": {"label": "Bedouin Space", "color": "#ef4444"},
        },
        "zones": zones,
        "stalls": uniq,
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w") as f:
        json.dump(out, f, indent=1)
    print(f"wrote {len(uniq)} stalls -> {os.path.relpath(OUT)}")
    print("capacity:", caps)
    print("zones:", len(zones))
    print("sample:", uniq[0], uniq[47] if len(uniq) > 47 else "")


if __name__ == "__main__":
    main()
